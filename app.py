#!/usr/bin/env python3
"""
AC Cost Web — servidor Flask
Puerto: 5656
"""

import io, csv, os, json, re, time, requests
from collections import defaultdict
from datetime import datetime, date, time as dtime, timedelta
from zoneinfo import ZoneInfo
from flask import Flask, jsonify, render_template, request

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

app = Flask(__name__)

TZ              = ZoneInfo("Europe/Madrid")
ESIOS_TOKEN     = os.environ.get("ESIOS_TOKEN", "")
AEMET_TOKEN     = os.environ.get("AEMET_TOKEN", "")
ANTHROPIC_TOKEN = os.environ.get("ANTHROPIC_TOKEN", "")
ESIOS_BASE      = "https://api.esios.ree.es/indicators"
AEMET_BASE      = "https://opendata.aemet.es/opendata/api"
ESIOS_INDICATOR = 1001
GEO_PENINSULA   = 8741

CONFIG_FILE      = os.environ.get("CONFIG_FILE",      "/config/settings.json")
HISTORY_FILE     = os.environ.get("HISTORY_FILE",     "/config/history.json")
CUSTOM_DEV_FILE  = os.environ.get("CUSTOM_DEV_FILE",  "/config/custom_devices.json")

MACHINES = [
    {"id": "mitsubishi_grande",  "label": "Mitsubishi MSZ-HR35VF", "kw_frio": 1.21, "kw_calor": 0.975},
    {"id": "mitsubishi_pequena", "label": "Mitsubishi MSZ-HR25VF",  "kw_frio": 0.80, "kw_calor": 0.850},
    {"id": "lg_viejita",         "label": "LG AS-H126RKA2",          "kw_frio": 1.30, "kw_calor": 1.20},
]

DEFAULT_SETTINGS = {
    "aemet_municipio":        "08200",
    "aemet_municipio_nombre": "Sant Boi de Llobregat",
    "bono_social_pct":        42.5,
    "tarifa_defecto":         "pvpc",
}

# ── Caché ─────────────────────────────────────────────────────────────────────
_cache = {}
def cache_get(key):
    e = _cache.get(key)
    return e["val"] if e and time.time() < e["exp"] else None
def cache_set(key, val, ttl=3600):
    _cache[key] = {"val": val, "exp": time.time() + ttl}

# ── Settings ──────────────────────────────────────────────────────────────────
def load_settings():
    try:
        with open(CONFIG_FILE) as f:
            return {**DEFAULT_SETTINGS, **json.load(f)}
    except Exception:
        return DEFAULT_SETTINGS.copy()

def save_settings(data):
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    current = load_settings()
    current.update(data)
    with open(CONFIG_FILE, "w") as f:
        json.dump(current, f, indent=2, ensure_ascii=False)
    return current

# ── Historial ─────────────────────────────────────────────────────────────────
def load_history():
    try:
        with open(HISTORY_FILE) as f:
            return json.load(f)
    except Exception:
        return []

def save_history(entry):
    os.makedirs(os.path.dirname(HISTORY_FILE), exist_ok=True)
    hist = load_history()
    hist.insert(0, entry)
    hist = hist[:50]
    with open(HISTORY_FILE, "w") as f:
        json.dump(hist, f, indent=2, ensure_ascii=False)
    return hist

# ── Aparatos personalizados ───────────────────────────────────────────────────
def load_custom_devices():
    try:
        with open(CUSTOM_DEV_FILE) as f:
            return json.load(f)
    except Exception:
        return []

def save_custom_device(device):
    os.makedirs(os.path.dirname(CUSTOM_DEV_FILE), exist_ok=True)
    devices = load_custom_devices()
    # Actualizar si ya existe el mismo id
    devices = [d for d in devices if d.get("id") != device.get("id")]
    devices.append(device)
    with open(CUSTOM_DEV_FILE, "w") as f:
        json.dump(devices, f, indent=2, ensure_ascii=False)
    return devices

def delete_custom_device(device_id):
    os.makedirs(os.path.dirname(CUSTOM_DEV_FILE), exist_ok=True)
    devices = [d for d in load_custom_devices() if d.get("id") != device_id]
    with open(CUSTOM_DEV_FILE, "w") as f:
        json.dump(devices, f, indent=2, ensure_ascii=False)
    return devices

# ── Rutas ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    settings       = load_settings()
    custom_devices = load_custom_devices()
    return render_template("index.html", machines=MACHINES,
                           settings=settings, custom_devices=custom_devices)

@app.route("/health")
def health():
    return jsonify({"status": "ok", "esios": bool(ESIOS_TOKEN),
                    "aemet": bool(AEMET_TOKEN), "ai": bool(ANTHROPIC_TOKEN)})

# ── Settings API ──────────────────────────────────────────────────────────────
@app.route("/api/settings", methods=["GET"])
def get_settings():
    return jsonify(load_settings())

@app.route("/api/settings", methods=["POST"])
def post_settings():
    return jsonify(save_settings(request.json or {}))

# ── Historial API ─────────────────────────────────────────────────────────────
@app.route("/api/history", methods=["GET"])
def get_history():
    return jsonify(load_history())

@app.route("/api/history", methods=["POST"])
def post_history():
    entry = request.json or {}
    entry["ts"] = datetime.now(TZ).isoformat()
    return jsonify(save_history(entry))

@app.route("/api/history/<int:idx>", methods=["DELETE"])
def delete_history_entry(idx):
    hist = load_history()
    if 0 <= idx < len(hist):
        hist.pop(idx)
        os.makedirs(os.path.dirname(HISTORY_FILE), exist_ok=True)
        with open(HISTORY_FILE, "w") as f:
            json.dump(hist, f, indent=2, ensure_ascii=False)
    return jsonify({"ok": True})

@app.route("/api/history", methods=["DELETE"])
def clear_history():
    os.makedirs(os.path.dirname(HISTORY_FILE), exist_ok=True)
    with open(HISTORY_FILE, "w") as f:
        json.dump([], f)
    return jsonify({"ok": True})

# ── Aparatos personalizados API ───────────────────────────────────────────────
@app.route("/api/custom-devices", methods=["GET"])
def get_custom_devices():
    return jsonify(load_custom_devices())

@app.route("/api/custom-devices", methods=["POST"])
def post_custom_device():
    device = request.json or {}
    if not device.get("id") or not device.get("label"):
        return jsonify({"error": "Faltan campos id y label"}), 400
    return jsonify(save_custom_device(device))

@app.route("/api/custom-devices/<device_id>", methods=["DELETE"])
def del_custom_device(device_id):
    return jsonify(delete_custom_device(device_id))

# ── IA: buscar consumo de aparato ─────────────────────────────────────────────
@app.route("/api/search-consumption")
def search_consumption():
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify({"error": "Falta el parámetro q"}), 400
    if not ANTHROPIC_TOKEN:
        return jsonify({"error": "ANTHROPIC_TOKEN no configurado"}), 500

    cached = cache_get(f"consumption:{query.lower()}")
    if cached:
        return jsonify(cached)

    prompt = f"""Eres un experto en eficiencia energética. El usuario quiere saber el consumo eléctrico típico de este aparato: "{query}".

Responde ÚNICAMENTE con un JSON válido, sin texto adicional, sin bloques de código, sin explicaciones:
{{
  "label": "nombre normalizado del aparato en español",
  "kw_min": número decimal (consumo mínimo en kW),
  "kw_max": número decimal (consumo máximo en kW),
  "kw_tipico": número decimal (consumo típico/recomendado en kW),
  "nota": "una frase breve explicando el rango o el modo de uso típico"
}}

Ejemplos de valores correctos:
- Televisor 55" OLED: kw_min=0.05, kw_max=0.15, kw_tipico=0.10
- Lavadora ciclo normal: kw_min=0.5, kw_max=2.5, kw_tipico=1.0
- Horno eléctrico: kw_min=1.0, kw_max=3.5, kw_tipico=2.2

Si no reconoces el aparato, devuelve kw_tipico=null y nota explicando que no tienes datos."""

    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_TOKEN,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 300,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=20,
        )
        r.raise_for_status()
        text = r.json()["content"][0]["text"].strip()
        # Limpiar bloques markdown ```json ... ``` si Claude los incluye
        text = re.sub(r"^```[a-z]*\n?", "", text)
        text = re.sub(r"```$", "", text).strip()
        result = json.loads(text)
        cache_set(f"consumption:{query.lower()}", result, ttl=86400)
        return jsonify(result)
    except json.JSONDecodeError:
        return jsonify({"error": "La IA no devolvió JSON válido", "raw": text[:200]}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── Utilidades para parsear ficheros de consumo ─────────────────────────────────
def normalize_header(value):
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = s.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ü', 'u').replace('ñ', 'n')
    s = re.sub(r'[^a-z0-9]+', '_', s)
    return s.strip('_')

def parse_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(',', '.')
    s = re.sub(r'[^0-9\.-]+', '', s)
    if not s or s in ['.', '-', '-.']:
        return None
    try:
        return float(s)
    except ValueError:
        return None

def parse_date_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y', '%Y.%m.%d']:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def parse_time_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, dtime):
        return value
    if isinstance(value, datetime):
        return value.time()
    s = str(value).strip()
    for fmt in ['%H:%M', '%H.%M', '%H%M', '%H']:
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    if '-' in s:
        parts = [p.strip() for p in s.split('-') if p.strip()]
        if parts:
            return parse_time_value(parts[0])
    return None

def parse_hour_label(value):
    """Devuelve (hora_del_dia 0-23, desplazamiento_dias) a partir de la columna
    'Hora'. Soporta rangos 'HH:MM-HH:MM' (se coge el inicio) y la convención de
    etiqueta numérica 1-24 que usan algunos exportadores de distribuidora,
    donde la etiqueta 24 representa las 23:00-24:00 del día ANTERIOR a la
    fecha que aparece en esa fila del fichero (hay que restar un día)."""
    if value is None or value == "":
        return None, 0
    s = str(value).strip()
    if ':' in s:
        t = parse_time_value(s)
        return (t.hour, 0) if t else (None, 0)
    n = parse_number(s)
    if n is not None and 1 <= n <= 24:
        label = int(round(n))
        if label == 24:
            return 23, -1
        if 1 <= label <= 23:
            return label - 1, 0
    t = parse_time_value(s)
    return (t.hour, 0) if t else (None, 0)

def read_spreadsheet(file_storage):
    filename = file_storage.filename or ''
    ext = os.path.splitext(filename)[1].lower()
    content = file_storage.read()
    file_storage.seek(0)
    if ext in ('.csv', '.txt'):
        text = content.decode('utf-8', errors='replace')
        try:
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=';,')
        except Exception:
            dialect = csv.excel
        rows = [row for row in csv.reader(text.splitlines(), dialect) if any(cell.strip() for cell in row)]
        return rows

    if ext == '.xls' and xlrd:
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        rows = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                ctype = sheet.cell_type(r, c)
                value = sheet.cell_value(r, c)
                if ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = datetime(*xlrd.xldate_as_tuple(value, book.datemode))
                    except Exception:
                        pass
                row.append(value)
            rows.append(row)
        return rows

    if ext in ('.xlsx', '.xlsm', '.xlsb') and load_workbook:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
        sheet = wb.active
        rows = [[cell.value for cell in row] for row in sheet.iter_rows(values_only=True)]
        return rows

    if load_workbook:
        try:
            wb = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
            sheet = wb.active
            rows = [[cell.value for cell in row] for row in sheet.iter_rows(values_only=True)]
            return rows
        except Exception:
            pass
    if xlrd:
        try:
            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_index(0)
            rows = []
            for r in range(sheet.nrows):
                row = []
                for c in range(sheet.ncols):
                    ctype = sheet.cell_type(r, c)
                    value = sheet.cell_value(r, c)
                    if ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = datetime(*xlrd.xldate_as_tuple(value, book.datemode))
                        except Exception:
                            pass
                    row.append(value)
                rows.append(row)
            return rows
        except Exception:
            pass
    raise ValueError('Formato de fichero no soportado. Usa XLS, XLSX o CSV.')

def parse_consumption_rows(rows):
    if not rows or len(rows) < 2:
        raise ValueError('El fichero debe contener una fila de cabecera y al menos una fila de datos.')

    # Algunos exportadores (p.ej. distribuidoras eléctricas) anteponen filas de
    # metadatos (CUPS, fechas del periodo...) antes de la fila de cabecera real.
    # Buscamos la primera fila que parezca cabecera (con columna de fecha y de hora).
    def is_date_col(h):
        return h in ('fecha', 'date', 'dia', 'day') or h.startswith('fecha_') or h.startswith('date_')

    def is_hour_col(h):
        return h in ('hora', 'time', 'periodo', 'period', 'hour') or h.startswith('hora_') or h.startswith('periodo_')

    header_row_idx = None
    headers = None
    for i, row in enumerate(rows[:15]):
        candidate = [normalize_header(v) for v in row]
        if not any(candidate):
            continue
        has_date = any(h and is_date_col(h) for h in candidate)
        has_hour = any(h and is_hour_col(h) for h in candidate)
        has_value = any(h and ('consumo' in h or 'kwh' in h or 'energia' in h or 'valor' in h) for h in candidate)
        if has_date and has_hour and has_value:
            header_row_idx = i
            headers = candidate
            break
    if header_row_idx is None:
        for i, row in enumerate(rows[:15]):
            candidate = [normalize_header(v) for v in row]
            if not any(candidate):
                continue
            has_date = any(h and is_date_col(h) for h in candidate)
            has_hour = any(h and is_hour_col(h) for h in candidate)
            if has_date and has_hour:
                header_row_idx = i
                headers = candidate
                break
    if header_row_idx is None:
        header_row_idx = 0
        headers = [normalize_header(v) for v in rows[0]]
    if not any(headers):
        raise ValueError('No se ha encontrado una cabecera válida en el fichero.')

    def find_header(keys):
        for key in keys:
            for idx, h in enumerate(headers):
                if h and key in h:
                    return idx
        return None

    datetime_idx = find_header(['fecha_hora', 'datetime', 'timestamp', 'fecha_y_hora', 'fecha_y_hora', 'fecha_y_hora', 'fecha_y_hora'])
    date_idx = find_header(['fecha', 'date', 'dia', 'day'])
    time_idx = find_header(['hora', 'time', 'periodo', 'period', 'hour'])
    value_idx = find_header(['consumo', 'kwh', 'energia', 'energy', 'usage', 'valor'])
    power_idx = find_header(['kw', 'potencia', 'power'])
    duration_idx = find_header(['horas', 'duration', 'hours'])
    cost_idx = find_header(['coste_por_hora', 'importe', 'coste'])

    # Detecta la unidad de la columna de consumo (Wh vs kWh) para normalizar a kWh.
    value_unit_factor = 1.0
    if value_idx is not None:
        vh = headers[value_idx]
        if 'kwh' in vh:
            value_unit_factor = 1.0
        elif 'wh' in vh:
            value_unit_factor = 0.001

    records = []
    for row in rows[header_row_idx + 1:]:
        if not any(cell not in (None, '') for cell in row):
            continue
        dt = None
        if datetime_idx is not None and datetime_idx < len(row):
            raw = row[datetime_idx]
            if isinstance(raw, datetime):
                dt = raw
            elif isinstance(raw, (int, float)):
                if xlrd and isinstance(raw, float):
                    try:
                        dt = datetime(*xlrd.xldate_as_tuple(raw, 0))
                    except Exception:
                        dt = None
            elif isinstance(raw, str) and raw.strip():
                try:
                    dt = datetime.fromisoformat(raw.strip())
                except Exception:
                    for fmt in ['%d/%m/%Y %H:%M', '%d-%m-%Y %H:%M', '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S']:
                        try:
                            dt = datetime.strptime(raw.strip(), fmt)
                            break
                        except ValueError:
                            continue
        row_date = None
        if date_idx is not None and date_idx < len(row):
            row_date = parse_date_value(row[date_idx])
        row_hour, day_offset = (None, 0)
        if time_idx is not None and time_idx < len(row):
            row_hour, day_offset = parse_hour_label(row[time_idx])
        if row_date is not None and day_offset:
            row_date = row_date + timedelta(days=day_offset)
        if dt is None and row_date and row_hour is not None:
            dt = datetime.combine(row_date, dtime(row_hour, 0))
        if dt is None and row_date:
            dt = datetime.combine(row_date, dtime(0, 0))

        consumption = None
        if value_idx is not None and value_idx < len(row):
            consumption = parse_number(row[value_idx])
            if consumption is not None:
                consumption *= value_unit_factor
        if consumption is None and power_idx is not None and power_idx < len(row):
            power = parse_number(row[power_idx])
            duration = None
            if duration_idx is not None and duration_idx < len(row):
                duration = parse_number(row[duration_idx])
            if power is not None:
                consumption = power * (duration if duration is not None else 1.0)
        if consumption is None:
            for cell in row:
                num = parse_number(cell)
                if num is not None:
                    consumption = num
                    break
        if consumption is None:
            continue

        record_date = dt.date() if isinstance(dt, datetime) else row_date
        if record_date is None:
            continue

        record_hour = dt.hour if isinstance(dt, datetime) else row_hour

        record_cost = None
        if cost_idx is not None and cost_idx < len(row):
            record_cost = parse_number(row[cost_idx])

        records.append({
            'datetime': dt.isoformat() if isinstance(dt, datetime) else None,
            'date': record_date.isoformat(),
            'hour': record_hour,
            'kwh': round(consumption, 6),
            'cost': record_cost,
        })

    if not records:
        raise ValueError('No se han podido extraer filas de consumo del fichero.')
    return records

@app.route('/api/consumption/upload', methods=['POST'])
def upload_consumption():
    if 'file' not in request.files:
        return jsonify({'error': 'Falta el fichero en la petición.'}), 400
    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'error': 'Falta el nombre del fichero.'}), 400
    try:
        rows = read_spreadsheet(f)
        records = parse_consumption_rows(rows)
        dates = sorted({r['date'] for r in records})

        # Si el propio fichero trae el coste horario ya calculado por la
        # distribuidora (columna "Coste por hora"), lo usamos directamente:
        # es más preciso que reconstruirlo con el precio PVPC de ESIOS (evita
        # desajustes de redondeo/alineación horaria) y no requiere ESIOS_TOKEN.
        # Algunos exportadores rellenan esa columna a 0.0 para todas las filas
        # (sin datos de precio reales) — en ese caso caemos a ESIOS.
        file_has_cost = any(r.get('cost') not in (None, 0, 0.0) for r in records)

        price_cache = {}
        if not file_has_cost:
            if not ESIOS_TOKEN:
                return jsonify({'error': 'ESIOS_TOKEN no configurado en .env'}), 500
            missing = []
            for date_str in dates:
                data = _fetch_esios(date_str)
                if not data:
                    missing.append(date_str)
                else:
                    price_cache[date_str] = data
            if missing:
                return jsonify({'error': f'No hay datos de precio ESIOS para estas fechas: {", ".join(missing)}'}), 400

        total_kwh = 0.0
        total_cost = 0.0
        by_date = defaultdict(lambda: {'kwh': 0.0, 'cost': 0.0, 'avg_price': 0.0, 'rows': 0})
        detail = []
        for rec in records:
            if file_has_cost:
                cost = rec['cost'] or 0.0
                price = cost / rec['kwh'] if rec['kwh'] else 0.0
            else:
                price_data = price_cache.get(rec['date'])
                price = price_data['hours'].get(rec['hour']) if rec['hour'] is not None else price_data['avg']
                if price is None:
                    price = price_data['avg']
                cost = rec['kwh'] * price
            total_kwh += rec['kwh']
            total_cost += cost
            entry = by_date[rec['date']]
            entry['kwh'] += rec['kwh']
            entry['cost'] += cost
            entry['rows'] += 1
            entry['avg_price'] += price
            if len(detail) < 20:
                detail.append({
                    'datetime': rec['datetime'] or rec['date'],
                    'kwh': rec['kwh'],
                    'price': price,
                    'cost': round(cost, 6),
                })
        dates_summary = []
        for date_str in dates:
            entry = by_date[date_str]
            if entry['rows']:
                avg_price = entry['avg_price'] / entry['rows']
            elif entry['kwh']:
                avg_price = entry['cost'] / entry['kwh']
            else:
                avg_price = 0
            dates_summary.append({
                'date': date_str,
                'kwh': round(entry['kwh'], 6),
                'cost': round(entry['cost'], 6),
                'avg_price': round(avg_price, 6),
            })

        return jsonify({
            'file_name': f.filename,
            'period_start': dates[0],
            'period_end': dates[-1],
            'rows': len(records),
            'total_kwh': round(total_kwh, 6),
            'total_cost': round(total_cost, 6),
            'avg_price': round(total_cost / total_kwh, 6) if total_kwh else 0,
            'dates': dates_summary,
            'detail': detail,
            'price_source': 'fichero' if file_has_cost else 'esios',
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── ESIOS: precios ────────────────────────────────────────────────────────────
def _fetch_esios(date_str):
    cached = cache_get(f"prices:{date_str}")
    if cached:
        return cached
    r = requests.get(f"{ESIOS_BASE}/{ESIOS_INDICATOR}",
        params={"start_date": f"{date_str}T00:00:00+02:00",
                "end_date":   f"{date_str}T23:59:59+02:00"},
        headers={"Accept": "application/json; application/vnd.esios-api-v1+json",
                 "Content-Type": "application/json",
                 "x-api-key": ESIOS_TOKEN}, timeout=15)
    r.raise_for_status()
    peninsula = [v for v in r.json().get("indicator", {}).get("values", [])
                 if v.get("geo_id") == GEO_PENINSULA]
    if not peninsula:
        return None
    ph = {}
    for v in peninsula:
        dt = datetime.fromisoformat(v.get("datetime","").replace("Z","+00:00")).astimezone(TZ)
        ph[dt.hour] = round(v["value"] / 1000, 6)
    result = {"date": date_str, "hours": ph,
              "avg": round(sum(ph.values())/len(ph), 6),
              "min": round(min(ph.values()), 6), "max": round(max(ph.values()), 6),
              "min_h": min(ph, key=ph.get), "max_h": max(ph, key=ph.get)}
    ttl = 1800 if date_str == date.today().isoformat() else 86400
    cache_set(f"prices:{date_str}", result, ttl=ttl)
    return result

@app.route("/api/prices")
def prices():
    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "Falta date (YYYY-MM-DD)"}), 400
    if not ESIOS_TOKEN:
        return jsonify({"error": "ESIOS_TOKEN no configurado en .env"}), 500
    try:
        result = _fetch_esios(date_str)
        if not result:
            return jsonify({"error": f"Sin datos para {date_str} (se publican ~20:30h del día anterior)"}), 404
        return jsonify(result)
    except requests.exceptions.HTTPError as e:
        return jsonify({"error": f"ESIOS HTTP {e.response.status_code}"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── ESIOS: histórico 30 días ──────────────────────────────────────────────────
@app.route("/api/prices/history")
def prices_history():
    if not ESIOS_TOKEN:
        return jsonify({"error": "ESIOS_TOKEN no configurado"}), 500
    cached = cache_get("prices_history_30d")
    if cached:
        return jsonify(cached)
    today = date.today()
    result = {}
    for delta in range(30):
        d = today - timedelta(days=delta)
        date_str = d.isoformat()
        try:
            data = _fetch_esios(date_str)
            if data:
                result[date_str] = {"avg": data["avg"], "min": data["min"], "max": data["max"]}
        except Exception:
            pass
    cache_set("prices_history_30d", result, ttl=3600)
    return jsonify(result)

# ── AEMET: municipios ─────────────────────────────────────────────────────────
@app.route("/api/municipios")
def municipios():
    q = request.args.get("q", "").strip().lower()
    if not q or len(q) < 3:
        return jsonify({"error": "Introduce al menos 3 caracteres"}), 400
    if not AEMET_TOKEN:
        return jsonify({"error": "AEMET_TOKEN no configurado"}), 500
    cached = cache_get("municipios_list")
    if not cached:
        try:
            r = requests.get(f"{AEMET_BASE}/maestro/municipios",
                headers={"api_key": AEMET_TOKEN, "Accept": "application/json"}, timeout=15)
            r.raise_for_status()
            r2 = requests.get(r.json().get("datos"), timeout=15)
            r2.raise_for_status()
            cached = r2.json()
            cache_set("municipios_list", cached, ttl=86400)
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    results = [{"id": m["id"].replace("id",""), "nombre": m["nombre"]}
               for m in cached if q in m.get("nombre","").lower()][:10]
    return jsonify(results)

# ── AEMET: temperatura horaria ────────────────────────────────────────────────
@app.route("/api/temperature")
def temperature():
    date_str  = request.args.get("date")
    settings  = load_settings()
    municipio = request.args.get("municipio", settings["aemet_municipio"])
    if not date_str:
        return jsonify({"error": "Falta date"}), 400
    if not AEMET_TOKEN:
        return jsonify({"error": "AEMET_TOKEN no configurado"}), 500
    cached = cache_get(f"temp:{date_str}:{municipio}")
    if cached:
        return jsonify(cached)
    try:
        r1 = requests.get(f"{AEMET_BASE}/prediccion/especifica/municipio/horaria/{municipio}",
            headers={"api_key": AEMET_TOKEN, "Accept": "application/json"}, timeout=15)
        if r1.status_code == 429:
            return jsonify({"error": "AEMET: demasiadas peticiones — espera unos minutos"}), 429
        r1.raise_for_status()
        data_url = r1.json().get("datos")
        if not data_url:
            return jsonify({"error": f"AEMET no devolvió datos para municipio {municipio}"}), 502
        r2 = requests.get(data_url, timeout=15)
        r2.raise_for_status()
        prediccion = json.loads(r2.content.decode("latin-1"))
        temps = {}
        for mun in prediccion:
            for dia in mun.get("prediccion", {}).get("dia", []):
                if dia.get("fecha", "")[:10] != date_str:
                    continue
                for t in dia.get("temperatura", []):
                    try:
                        periodo = str(t.get("periodo", "")).strip()
                        if len(periodo) == 2 and periodo.isdigit():
                            h = int(periodo)
                            if 0 <= h <= 23:
                                temps[h] = float(t.get("value", 0))
                    except (ValueError, TypeError):
                        pass
        if not temps:
            return jsonify({"error": "AEMET solo publica horas futuras para hoy. Prueba mañana."}), 404
        result = {"date": date_str, "municipio": municipio, "temps": temps,
                  "min": min(temps.values()), "max": max(temps.values())}
        cache_set(f"temp:{date_str}:{municipio}", result, ttl=21600)
        return jsonify(result)
    except requests.exceptions.HTTPError as e:
        code = e.response.status_code
        if code == 429:
            return jsonify({"error": "AEMET: demasiadas peticiones"}), 429
        return jsonify({"error": f"AEMET HTTP {code}"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── Precios medios semanales ──────────────────────────────────────────────────
@app.route("/api/prices/weekly-avg")
def weekly_avg():
    if not ESIOS_TOKEN:
        return jsonify({"error": "ESIOS_TOKEN no configurado"}), 500
    cached = cache_get("weekly_avg")
    if cached:
        return jsonify(cached)
    today = date.today()
    avgs  = {}
    for delta in range(28):
        d = today - timedelta(days=delta+1)
        dow = d.weekday()
        try:
            data = _fetch_esios(d.isoformat())
            if data:
                if dow not in avgs:
                    avgs[dow] = {}
                for h, p in data["hours"].items():
                    k = str(h)
                    if k not in avgs[dow]:
                        avgs[dow][k] = []
                    avgs[dow][k].append(float(p))
        except Exception:
            pass
    result = {dow: {h: round(sum(ps)/len(ps), 6) for h, ps in hours.items()}
              for dow, hours in avgs.items()}
    cache_set("weekly_avg", result, ttl=3600)
    return jsonify(result)

@app.route("/api/machines")
def machines_route():
    return jsonify(MACHINES)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5656))
    app.run(host="0.0.0.0", port=port, debug=False)
